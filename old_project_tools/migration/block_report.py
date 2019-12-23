# -*- coding: utf-8 -*-
from pymongo import MongoClient
from collections import defaultdict
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


main_db_host = 'srv-5.yottos.com:27018,srv-5.yottos.com:27019,srv-5.yottos.com:27020'

conn = MongoClient(host=main_db_host)
db = conn.getmyad_db

sites = {}
for item in db.domain.find({}, {'login': True, 'domains': True}):
    login = item.get('login')
    for k, v in item.get('domains', {}).iteritems():
        sites[k] = (login, v)

date = datetime.datetime.now()
date = datetime.datetime(date.year, date.month, date.day, 0, 0)

condition1 = {'date': {'$gte': date - datetime.timedelta(days=30), '$lt': date}}
condition2 = {'date': {'$gte': date - datetime.timedelta(days=90), '$lt': date}}
reduce = '''
        function(o, p) {
           p.totalCost += o.totalCost || 0;
           p.impressions_block += o.impressions_block || 0;
           p.impressions_block_not_valid += o.impressions_block_not_valid || 0;
           p.impressions += o.impressions || 0;
           p.clicks += o.clicks || 0;
           p.clicksUnique += o.clicksUnique || 0;
           p.social_impressions += o.social_impressions || 0;
           p.social_clicks += o.social_clicks || 0;
           p.social_clicksUnique += o.social_clicksUnique || 0;
        }'''
initial = {'totalCost': 0,
           'impressions_block': 0,
           'impressions_block_not_valid': 0,
           'impressions': 0,
           'clicks': 0,
           'clicksUnique': 0,
           'social_impressions': 0,
           'social_clicks': 0,
           'social_clicksUnique': 0
           }
cur1 = db.stats.daily.domain.group(
    key=['domain_guid'],
    condition=condition1,
    reduce=reduce,
    initial=initial
)
cur2 = db.stats.daily.domain.group(
    key=['domain_guid'],
    condition=condition2,
    reduce=reduce,
    initial=initial
)

stats = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

for item in cur1:
    stats[item['domain_guid']]['30'] = {
        'totalCost': item['totalCost'],
        'impressions_block': item['impressions_block'],
        'impressions_block_not_valid': item['impressions_block_not_valid'],
        'impressions': item['impressions'],
        'clicks': item['clicks'],
        'clicksUnique': item['clicksUnique'],
        'social_impressions': item['social_impressions'],
        'social_clicks': item['social_clicks'],
        'social_clicksUnique': item['social_clicksUnique'],
    }

for item in cur2:
    stats[item['domain_guid']]['90'] = {
        'totalCost': item['totalCost'],
        'impressions_block': item['impressions_block'],
        'impressions_block_not_valid': item['impressions_block_not_valid'],
        'impressions': item['impressions'],
        'clicks': item['clicks'],
        'clicksUnique': item['clicksUnique'],
        'social_impressions': item['social_impressions'],
        'social_clicks': item['social_clicks'],
        'social_clicksUnique': item['social_clicksUnique'],
    }

wb = Workbook()
ws = wb.get_active_sheet()
ws.cell(row=1, column=1).value = 'Логин'
ws.cell(row=1, column=2).value = 'Сайт'
ws.cell(row=1, column=3).value = '30 дней'
ws.cell(row=2, column=3).value = 'Сумма'
ws.cell(row=2, column=4).value = 'Показы блока'
ws.cell(row=2, column=5).value = 'Гарантированные показы блока'
ws.cell(row=2, column=6).value = 'Показы РП'
ws.cell(row=2, column=7).value = 'Клики'
ws.cell(row=2, column=8).value = 'Уникальные клики'
ws.cell(row=2, column=9).value = 'Социальные показы'
ws.cell(row=2, column=10).value = 'Социальные клики'
ws.cell(row=2, column=11).value = 'Социальные уникальные клики'
ws.cell(row=1, column=12).value = '90 дней'
ws.cell(row=2, column=12).value = 'Сумма'
ws.cell(row=2, column=13).value = 'Показы блока'
ws.cell(row=2, column=14).value = 'Гарантированные показы блока'
ws.cell(row=2, column=15).value = 'Показы РП'
ws.cell(row=2, column=16).value = 'Клики'
ws.cell(row=2, column=17).value = 'Уникальные клики'
ws.cell(row=2, column=18).value = 'Социальные показы'
ws.cell(row=2, column=19).value = 'Социальные клики'
ws.cell(row=2, column=20).value = 'Социальные уникальные клики'
ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=11)
ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=20)
row_count = 3
for k, v in stats.iteritems():
    acc, site = sites.get(k, ('NOT FOUND', k))
    try:
        del sites[k]
    except Exception:
        pass
    ws.cell(row=row_count, column=1).value = acc
    ws.cell(row=row_count, column=2).value = site
    ws.cell(row=row_count, column=3).value = v['30']['totalCost']
    ws.cell(row=row_count, column=4).value = v['30']['impressions_block']
    ws.cell(row=row_count, column=5).value = v['30']['impressions_block_not_valid']
    ws.cell(row=row_count, column=6).value = v['30']['impressions']
    ws.cell(row=row_count, column=7).value = v['30']['clicks']
    ws.cell(row=row_count, column=8).value = v['30']['clicksUnique']
    ws.cell(row=row_count, column=9).value = v['30']['social_impressions']
    ws.cell(row=row_count, column=10).value = v['30']['social_clicks']
    ws.cell(row=row_count, column=11).value = v['30']['social_clicksUnique']
    ws.cell(row=row_count, column=12).value = v['90']['totalCost']
    ws.cell(row=row_count, column=13).value = v['90']['impressions_block']
    ws.cell(row=row_count, column=14).value = v['90']['impressions_block_not_valid']
    ws.cell(row=row_count, column=15).value = v['90']['impressions']
    ws.cell(row=row_count, column=16).value = v['90']['clicks']
    ws.cell(row=row_count, column=17).value = v['90']['clicksUnique']
    ws.cell(row=row_count, column=18).value = v['90']['social_impressions']
    ws.cell(row=row_count, column=19).value = v['90']['social_clicks']
    ws.cell(row=row_count, column=20).value = v['90']['social_clicksUnique']
    row_count += 1

for acc, site in sites.itervalues():
    ws.cell(row=row_count, column=1).value = acc
    ws.cell(row=row_count, column=2).value = site
    ws.cell(row=row_count, column=3).value = 0
    ws.cell(row=row_count, column=4).value = 0
    ws.cell(row=row_count, column=5).value = 0
    ws.cell(row=row_count, column=6).value = 0
    ws.cell(row=row_count, column=7).value = 0
    ws.cell(row=row_count, column=8).value = 0
    ws.cell(row=row_count, column=9).value = 0
    ws.cell(row=row_count, column=10).value = 0
    ws.cell(row=row_count, column=11).value = 0
    ws.cell(row=row_count, column=12).value = 0
    ws.cell(row=row_count, column=13).value = 0
    ws.cell(row=row_count, column=14).value = 0
    ws.cell(row=row_count, column=15).value = 0
    ws.cell(row=row_count, column=16).value = 0
    ws.cell(row=row_count, column=17).value = 0
    ws.cell(row=row_count, column=18).value = 0
    ws.cell(row=row_count, column=19).value = 0
    ws.cell(row=row_count, column=20).value = 0
    row_count += 1

wb.save('./rep.xls')
