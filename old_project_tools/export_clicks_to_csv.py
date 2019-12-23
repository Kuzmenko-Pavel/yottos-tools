# -*- coding: utf-8 -*-
import os
import sys
import time

project_dir = os.path.dirname(os.path.realpath(__file__))
sys.path.append(project_dir)
os.environ['PYTHON_EGG_CACHE'] = '/usr/lib/python2.7/dist-packages'
import datetime
import sys
from pymongo import MongoClient
from collections import defaultdict
import urlparse
import urllib
import pymssql
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import StringIO
import ftplib
import requests
import urllib3.contrib.pyopenssl

urllib3.contrib.pyopenssl.inject_into_urllib3()

sys.stdout = sys.stderr


current_dir = os.path.dirname(os.path.abspath(__file__))
rep_dir = os.path.join(current_dir, 'export_click')
if not os.path.exists(rep_dir):
    os.makedirs(rep_dir)

main_db_host = 'srv-5.yottos.com:27018,srv-5.yottos.com:27019,srv-5.yottos.com:27020'

conn = MongoClient(host=main_db_host)
db = conn.getmyad_db


informers = defaultdict(lambda: 'NOT TITLE')
for item in db.informer.find({}, {'guid': True, 'domain': True}):
    informers[item.get('guid')] = item.get('domain')

campaigns = defaultdict(lambda: 'NOT TITLE')
for item in db.campaign.find({}, {'guid': True, 'title': True}):
    campaigns[item.get('guid')] = item.get('title')

for item in db.campaign.archive.find({}, {'guid': True, 'title': True}):
    campaigns[item.get('guid')] = item.get('title')

date_now = datetime.datetime.now()
for day_iter in range(0, 18):
    date = datetime.datetime(date_now.year, date_now.month, date_now.day, 0, 0)
    date_start = date -  datetime.timedelta(days=day_iter)
    date_end = date_start + datetime.timedelta(days=1)
    print('Report be %s - %s' % (date_start, date_end))
    cursor = db.clicks.find({'dt': {'$gte': date_start, '$lt': date_end}, 'account_id' : '3e23c722-7446-4723-8985-b2fe120ac3a2'}).sort('dt', 1)
    wb = Workbook()
    ws = wb.create_sheet('', 0)
    row_count = 1
    ws.cell(row=row_count, column=1).value = 'Время'
    ws.cell(row=row_count, column=2).value = 'Компании'
    ws.cell(row=row_count, column=3).value = 'Оффер'
    ws.cell(row=row_count, column=4).value = 'Рекламный блок'
    ws.cell(row=row_count, column=5).value = 'Сайт'
    ws.cell(row=row_count, column=6).value = 'Цена'
    ws.cell(row=row_count, column=7).value = 'IP'
    ws.cell(row=row_count, column=8).value = 'Ссылка'
    ws.cell(row=row_count, column=9).value = 'Браузер'
    ws.cell(row=row_count, column=10).value = 'Реферер'
    ws.cell(row=row_count, column=11).value = 'Кук'
    c1 = ws['A1']
    c1.font = Font(size=14, bold=True)
    c1.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c1.column].width = 30
    c2 = ws['B1']
    c2.font = Font(size=14, bold=True)
    c2.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c2.column].width = 30
    c3 = ws['C1']
    c3.font = Font(size=14, bold=True)
    c3.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c3.column].width = 38
    c4 = ws['D1']
    c4.font = Font(size=14, bold=True)
    c4.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c4.column].width = 38
    c5 = ws['E1']
    c5.font = Font(size=14, bold=True)
    c5.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c5.column].width = 30
    c6 = ws['F1']
    c6.font = Font(size=14, bold=True)
    c6.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c6.column].width = 15
    c7 = ws['G1']
    c7.font = Font(size=14, bold=True)
    c7.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c7.column].width = 20

    c8 = ws['H1']
    c8.font = Font(size=14, bold=True)
    c8.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c8.column].width = 30

    c9 = ws['I1']
    c9.font = Font(size=14, bold=True)
    c9.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c9.column].width = 30

    c10 = ws['J1']
    c10.font = Font(size=14, bold=True)
    c10.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c10.column].width = 30

    c11 = ws['K1']
    c11.font = Font(size=14, bold=True)
    c11.alignment = Alignment(shrink_to_fit=False, wrap_text=True)
    ws.column_dimensions[c10.column].width = 30

    ws.row_dimensions[1].height = 25
    for doc in cursor:
        dt = doc.get('dt')
        campaign = campaigns[doc.get('campaignId', '')]
        block = doc.get('inf', '')
        domain = informers[block]
        cost = doc.get('adload_cost', 0)
        ip = doc.get('ip', '')
        url = doc.get('url', '')
        user_agent = doc.get('user_agent', '')
        referer = doc.get('referer', '')
        cookie = doc.get('cookie', '')

        url_parts = list(urlparse.urlparse(url))
        query = dict(urlparse.parse_qsl(url_parts[4]))
        if query.has_key('utm_content'):
            offer = urllib.unquote(query['utm_content'].encode('utf-8'))
        else:
            offer = doc.get('offer', '')

        row_count += 1
        ws.cell(row=row_count, column=1).value = dt.strftime("%m/%d/%Y, %H:%M:%S")
        ws.cell(row=row_count, column=2).value = campaign
        ws.cell(row=row_count, column=3).value = offer
        ws.cell(row=row_count, column=4).value = block
        ws.cell(row=row_count, column=5).value = domain
        ws.cell(row=row_count, column=6).value = cost
        ws.cell(row=row_count, column=7).value = ip
        ws.cell(row=row_count, column=8).value = url
        ws.cell(row=row_count, column=9).value = user_agent
        ws.cell(row=row_count, column=10).value = referer
        ws.cell(row=row_count, column=11).value = cookie

    wb.save(os.path.join(rep_dir, '%s-%s-%s' % (date_start.year, date_start.month, date_start.day) + '.xlsx'))
