# -*- coding: utf-8 -*-
import sys

import os
import time

project_dir = os.path.dirname(os.path.realpath(__file__))
sys.path.append(project_dir)
os.environ['PYTHON_EGG_CACHE'] = '/usr/lib/python2.7/dist-packages'
import datetime
import sys
from pymongo import MongoClient
from collections import defaultdict
import urlparse
import urllib
import pymssql
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import StringIO
import ftplib
import requests
import urllib3.contrib.pyopenssl
urllib3.contrib.pyopenssl.inject_into_urllib3()


sys.stdout = sys.stderr


def chdir(ftp, dir):
    """

    Args:
        ftp:
        dir:
    """
    try:
        ftp.cwd(dir)
    except ftplib.all_errors as e:
        ftp.mkd(dir)
        print(e)
        ftp.cwd(dir)


def mssql_connection_adload():
    """

    Returns:

    """
    pymssql.set_max_connections(450)
    conn = pymssql.connect(host='srv-3.yottos.com',
                           user='web',
                           password='odif8duuisdofj',
                           database='AdLoad',
                           as_dict=True,
                           charset='cp1251')
    conn.autocommit(True)
    return conn


accounts = defaultdict(lambda: 'NOT TITLE')
connection_adload = mssql_connection_adload()
cursor = connection_adload.cursor()
cursor.execute('''SELECT lower([UserID]) as UserID, [Login] FROM [AdLoad].[dbo].[Users]''')
for row in cursor:
    accounts[row['UserID']] = row['Login'].decode('utf-8')
cursor.close()

main_db_host = 'srv-5.yottos.com:27018,srv-5.yottos.com:27019,srv-5.yottos.com:27020'

conn = MongoClient(host=main_db_host)
db = conn.getmyad_db
date = datetime.datetime.now()
date_start = datetime.datetime(date.year, date.month, date.day, 0, 0)
date_end = date_start - datetime.timedelta(days=7)
print('Report be %s - %s' % (date_end, date_start))

informers = defaultdict(lambda: 'NOT TITLE')
for item in db.informer.find({}, {'guid': True, 'domain': True}):
    informers[item.get('guid')] = item.get('domain')

campaigns = defaultdict(lambda: 'NOT TITLE')
campaigns_by_account = defaultdict(lambda: 'NOT_TITLE')
campaigns_by_hide_marker = defaultdict(lambda: True)
for item in db.campaign.find({}, {'guid': True, 'title': True, 'account': True, 'yottosHideSiteMarker': True}):
    campaigns[item.get('guid')] = item.get('title')
    campaigns_by_account[item.get('guid')] = item.get('account')
    campaigns_by_hide_marker[item.get('guid')] = item.get('yottosHideSiteMarker', True)

for item in db.campaign.archive.find({}, {'guid': True, 'title': True, 'account': True, 'yottosHideSiteMarker': True}):
    campaigns[item.get('guid')] = item.get('title')
    campaigns_by_account[item.get('guid')] = item.get('account')
    campaigns_by_hide_marker[item.get('guid')] = item.get('yottosHideSiteMarker', True)

pipeline = [
    {'$match':
        {
            'dt': {'$gte': date_end, '$lt': date_start},
            'adload_cost': {'$gt': 0}
        }
    },
    {'$group':
         {
             '_id': {
                 'campaign': '$campaignId',
                 'adv': '$inf',
                 'offer': '$offer',
                 'cost': '$adload_cost',
                 "day": {"$dayOfMonth": "$dt"},
                 "month": {"$month": "$dt"}
             },
             'url': {'$last': '$url'},
             'count': {
              '$sum': {"$cond": [
                  {"$eq": ["$unique", True]},
                  1,
                  0
              ]
              }
             }
,
             'summa': {
              '$sum': {"$cond": [
                  {"$eq": ["$unique", True]},
                  '$adload_cost',
                  0
              ]
              }
             }
          }
    },
    {'$sort': {"_id.day": 1}}
]
data = defaultdict(
    lambda: defaultdict(
        lambda: defaultdict(
            lambda: defaultdict(
                lambda: defaultdict(
                    lambda: defaultdict(
                        lambda: defaultdict(float)
                    )
                )
            )
        )
    )
)
cursor = db.clicks.aggregate(pipeline=pipeline, allowDiskUse=True, useCursor=True)
c = 0
for doc in cursor:
    date = '%s-%s' % (doc['_id']['month'], doc['_id']['day'])
    account = campaigns_by_account[doc['_id']['campaign']]
    campaign = campaigns[doc['_id']['campaign']]
    if campaigns_by_hide_marker[doc['_id']['campaign']]:
        block = doc['_id']['adv']
    else:
        block = informers[doc['_id']['adv']]

    url = doc['url']
    url_parts = list(urlparse.urlparse(url))
    query = dict(urlparse.parse_qsl(url_parts[4]))
    if query.has_key('utm_content'):
        offer = urllib.unquote(query['utm_content'].encode('utf-8'))
    else:
        offer = doc['_id']['offer']
    cost = doc['_id']['cost']
    count = doc['count']
    sum = doc['summa']
    data[account][date][campaign][offer][block][cost]['count'] = count
    data[account][date][campaign][offer][block][cost]['sum'] = sum
    # break


acl = []
for key, value in data.iteritems(): #account
    wb = Workbook()
    sheet_count = 0
    for k, v in value.iteritems(): #date
        ws = wb.create_sheet(k, sheet_count)
        row_count = 1
        ws.cell(row=row_count, column=1).value = 'Название компании'
        ws.cell(row=row_count, column=2).value = 'Id Тизера'
        ws.cell(row=row_count, column=3).value = 'Рекламный блок'
        ws.cell(row=row_count, column=4).value = 'Цена'
        ws.cell(row=row_count, column=5).value = 'Количество кликов'
        ws.cell(row=row_count, column=6).value = 'Сумма'
        c1 = ws['A1']
        c1.font = Font(size=14, bold=True)
        c1.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c1.column].width = 30
        c2 = ws['B1']
        c2.font = Font(size=14, bold=True)
        c2.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c2.column].width = 30
        c3 = ws['C1']
        c3.font = Font(size=14, bold=True)
        c3.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c3.column].width = 38
        c4 = ws['D1']
        c4.font = Font(size=14, bold=True)
        c4.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c4.column].width = 15
        c5 = ws['E1']
        c5.font = Font(size=14, bold=True)
        c5.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c5.column].width = 30
        c6 = ws['F1']
        c6.font = Font(size=14, bold=True)
        c6.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c6.column].width = 15
        ws.row_dimensions[1].height = 25
        for z, x in v.iteritems(): # campaign
            for a, s in x.iteritems(): #offer
                for q, w in s.iteritems(): #block
                    for e, r in w.iteritems():  # cost
                        row_count += 1
                        ws.cell(row=row_count, column=1).value = z
                        ws.cell(row=row_count, column=2).value = a
                        ws.cell(row=row_count, column=3).value = q
                        ws.cell(row=row_count, column=4).value = e
                        ws.cell(row=row_count, column=5).value = r['count']
                        ws.cell(row=row_count, column=6).value = r['sum']
        sheet_count += 1
    buf = StringIO.StringIO()
    wb.save(buf)
    buf.seek(0)
    ftp = ftplib.FTP(host='srv-10.yottos.com')
    ftp.login('cdn', '$www-app$')
    chdir(ftp, 'report')
    chdir(ftp, 'click')
    chdir(ftp, '%s-%s-%s' % (date_start.year, date_start.month, date_start.day))
    ftp.storbinary('STOR ' + str(key) + '_%s-%s-%s' % (date_start.year, date_start.month, date_start.day) + '.xlsx', buf)
    ftp.close()
    acl.append(key)

html_tr = ' \n'.join(['<tr><td><a href="%s">%s</a></td></tr>' % (
    'https://cdn.yottos.com/report/click/%s-%s-%s/%s%s.xlsx' % (
        date_start.year, date_start.month, date_start.day, x, '_%s-%s-%s' % (date_start.year, date_start.month, date_start.day)
    ), accounts[x]
) for x in acl])
html = '''
<html>
<head><title>Отчеты</title></head>
<body bgcolor="white">
<hr>
<center>
    <h3> Отчеты с %s по %s </h3>
</center>
<hr>
<center>
    <table>
        %s
    </table>
</center>
<hr>
</body>
</html>
''' % ('%s-%s-%s' % (date_end.year, date_end.month, date_end.day),
       '%s-%s-%s' % (date_start.year, date_start.month, date_start.day),
       html_tr.encode('utf-8'))
ftp = ftplib.FTP(host='srv-10.yottos.com')
ftp.login('cdn', '$www-app$')
chdir(ftp, 'report')
chdir(ftp, 'click')
ftp.storbinary('STOR index.html',  StringIO.StringIO(html))
ftp.close()

print('Report upload')
time.sleep(400)
headers = {'X-Cache-Update': '1'}
cdns = ['cdn.srv-10.yottos.com', 'cdn.srv-11.yottos.com', 'cdn.srv-12.yottos.com', 'cdn.yottos.com']

for cdn in cdns:
    for item in acl:
        url = 'http://%s%s' % (cdn, '/report/click/%s-%s-%s/%s%s.xlsx' % (
            date_start.year,
            date_start.month,
            date_start.day,
            item,
            '_%s-%s-%s' % (date_start.year, date_start.month, date_start.day)))
        r = requests.get(url, headers=headers, verify=False)
        print('%s - %s' % (url, r.status_code))
    url = 'http://%s%s' % (cdn, '/report/click/index.html')
    r = requests.get(url, headers=headers, verify=False)
    print('%s - %s' % (url, r.status_code))
print('Report create complite')