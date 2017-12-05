# This Python file uses the following encoding: utf-8
import sys

import os

project_dir = os.path.dirname(os.path.realpath(__file__))
sys.path.append(project_dir)
os.environ['PYTHON_EGG_CACHE'] = '/usr/lib/python2.7/dist-packages'
import datetime
from collections import defaultdict
import urllib
import urlparse
import sys
from pymongo import Connection
import xlwt
import StringIO
import ftplib
import collections
from openpyxl import Workbook


sys.stdout = sys.stderr

main_db_host = 'srv-5.yottos.com:27018,srv-5.yottos.com:27019,srv-5.yottos.com:27020'


conn = Connection(host=main_db_host)
db = conn.getmyad_db
informersBySite = {}
informersByItemsNumber = {}
informersByUsers = {}
informersByTitle = {}
informerList = []
for informer in db.informer.find({},
                                 {'guid': True, 'domain': True, 'admaker': True, 'user': True, 'title': True}):
    try:
        userGuid = db.users.find_one({"login": informer.get('user', 'NOT DOMAIN')}, {'guid': 1, '_id': 0})
        domainGuid = None
        for domains in db.domain.find({"login": informer.get('user', 'NOT DOMAIN')}, {'domains': 1, '_id': 0}):
            for key, value in domains['domains'].iteritems():
                if value == informer.get('domain', 'NOT DOMAIN'):
                    domainGuid = key
        informersBySite[informer['guid']] = (informer.get('domain', 'NOT DOMAIN'), domainGuid)
        informersByItemsNumber[informer['guid']] = informer.get('admaker', {}).get('Main', {}).get(
            'itemsNumber', 4)
        informersByUsers[informer['guid']] = (informer.get('user', 'NOT DOMAIN'), userGuid.get('guid', ''))
        informersByTitle[informer['guid']] = informer.get('title', 'NOT DOMAIN')
        informerList.append(informer['guid'])
    except:
        pass

campaigns = {}
for x in db.campaign.find():
    campaigns[x['guid']] = x['title']

for x in db.campaign.archive.find():
    campaigns[x['guid']] = x['title']

campaign_list = [x for x, y in campaigns.iteritems()]


pipeline = [
    {'$match':
         {
            'getmyad_user_id': '9b46cfc4-8f94-11e6-a0af-002590d97638'
         }
     },
    {'$group':
         {'_id': {'campaign': '$campaignId', 'adv': '$inf', "day":{"$dayOfMonth": "$dt"}, "month":{"$month": "$dt"}},
          'count_u': {
              '$sum': {"$cond": [
                  {"$eq": ["$unique", True]},
                  1,
                  0
              ]
              }
          },
          'count': {'$sum': 1}
        }
     },
    {'$sort': {"_id.day": 1}}
]
camp = {}
cursor = db.clicks.aggregate(pipeline=pipeline, cursor={})
stats = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
for doc in cursor:
    inf = informersBySite.get(doc['_id']['adv'], ['DELETED', ''])[0]
    d = "%s-%s" % (doc['_id']['month'], doc['_id']['day'])
    camp = campaigns.get(doc['_id']['campaign'], 'DELETED')
    count = stats[d][camp][inf]['count']
    count_u = stats[d][camp][inf]['count_u']
    stats[d][camp][inf]['count'] = count + doc['count']
    stats[d][camp][inf]['count_u'] = count + doc['count_u']


wb = Workbook()

for x, y in stats.iteritems():
    row = 0
    ws = wb.create_sheet(title=x)
    row += 1
    row_d = row
    sym_d = 0
    ws.cell(row=row, column=1).value = x
    for k, v in y.iteritems():
        row += 1
        row_c = row
        sym_c = 0
        ws.cell(row=row, column=2).value = k
        for z, x in v.iteritems():
            row += 1
            ws.cell(row=row, column=3).value = z
            ws.cell(row=row, column=4).value = x['count']
            ws.cell(row=row, column=5).value = x['count_u']
            sym_d += x['count']
            sym_c += x['count']
        ws.cell(row=row_c, column=3).value = sym_c
    ws.cell(row=row_d, column=2).value = sym_d
    row += 1
    row += 1

wb.save('visti_pro.xlsx')
