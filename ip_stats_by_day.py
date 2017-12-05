# This Python file uses the following encoding: utf-8
import sys

import os

project_dir = os.path.dirname(os.path.realpath(__file__))
sys.path.append(project_dir)
os.environ['PYTHON_EGG_CACHE'] = '/usr/lib/python2.7/dist-packages'
import datetime
import sys
from pymongo import Connection
from openpyxl import Workbook
from openpyxl.styles import PatternFill

sys.stdout = sys.stderr

main_db_host = 'srv-5.yottos.com:27018,srv-5.yottos.com:27019,srv-5.yottos.com:27020'

conn = Connection(host=main_db_host)
db = conn.getmyad_db

pipeline = [
    {'$match':
         {
            'dt': {'$gte': datetime.datetime(2017, 11, 1, 0, 0), '$lt': datetime.datetime(2017, 11, 15, 0, 0)},
         }
     },
    {'$group':
         {'_id': {'ip': '$ip', "day": {"$dayOfMonth": "$dt"}}, 'count': {'$sum': 1}}
     }
]

pipeline = [
    {'$match':
         {
            'dt': {'$gte': datetime.datetime(2017, 11, 1, 0, 0), '$lt': datetime.datetime(2017, 11, 15, 0, 0)},
         }
     },
    {'$group':
         {'_id': {'campaignId': '$campaignId'}, 'count': {'$sum': 1}}
     }
]
cursor = db.clicks.aggregate(pipeline=pipeline, cursor={})


for doc in cursor:
    print doc



# wb = Workbook()
# ws = wb.active
# ws.title = "ip_by_day"
#
# ws.cell(row=1, column=1).value = 'ip/date'
#
#
# for x in range(1, 15):
#     ws.cell(row=1, column=x+1).value = x
# row_count = 1
# ip_by_row = {}
# for doc in cursor:
#     count = int(doc.get('count', 0))
#     ip = doc.get('_id', {}).get('ip')
#     day = doc.get('_id', {}).get('day')
#     background = "FFFFFF"
#     if count > 0:
#         background = "05f943"
#     if ip is not None and day is not None:
#         row = ip_by_row.get(ip)
#         if row is None:
#             row_count += 1
#             row = row_count
#             ip_by_row[ip] = row
#             ws.cell(row=row, column=1).value = ip
#
#         ws.cell(row=row, column=day+1).value = count
#         ws.cell(row=row, column=day + 1).fill = PatternFill(bgColor=background, fill_type="solid")
#
#
# wb.save('ip_by_day.xlsx')
