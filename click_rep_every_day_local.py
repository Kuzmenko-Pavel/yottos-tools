# -*- coding: utf-8 -*-
import os
import sys
import time

project_dir = os.path.dirname(os.path.realpath(__file__))
sys.path.append(project_dir)
os.environ['PYTHON_EGG_CACHE'] = '/usr/lib/python2.7/dist-packages'
import datetime
import sys
from pymongo import MongoClient
from collections import defaultdict
import urlparse
import urllib
import pymssql
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import StringIO
import ftplib
import requests
import urllib3.contrib.pyopenssl

urllib3.contrib.pyopenssl.inject_into_urllib3()

sys.stdout = sys.stderr


def mssql_connection_adload():
    """

    Returns:

    """
    pymssql.set_max_connections(450)
    conn = pymssql.connect(host='srv-3.yottos.com',
                           user='web',
                           password='odif8duuisdofj',
                           database='AdLoad',
                           as_dict=True,
                           charset='cp1251')
    conn.autocommit(True)
    return conn


current_dir = os.path.dirname(os.path.abspath(__file__))
rep_dir = os.path.join(current_dir, 'report')
if not os.path.exists(rep_dir):
    os.makedirs(rep_dir)

accounts = defaultdict(lambda: 'NOT TITLE')
connection_adload = mssql_connection_adload()
cursor = connection_adload.cursor()
cursor.execute('''SELECT lower([UserID]) as UserID, [Login] FROM [AdLoad].[dbo].[Users]''')
for row in cursor:
    accounts[row['UserID']] = row['Login'].decode('utf-8')
cursor.close()

main_db_host = 'srv-5.yottos.com:27018,srv-5.yottos.com:27019,srv-5.yottos.com:27020'

conn = MongoClient(host=main_db_host)
db = conn.getmyad_db
date = datetime.datetime.now()
date_start = datetime.datetime(date.year, date.month, date.day, 0, 0)
date_end = date_start - datetime.timedelta(days=21)
print('Report be %s - %s' % (date_end, date_start))

informers = defaultdict(lambda: 'NOT TITLE')
for item in db.informer.find({}, {'guid': True, 'domain': True}):
    informers[item.get('guid')] = item.get('domain')

campaigns = defaultdict(lambda: 'NOT TITLE')
campaigns_by_account = defaultdict(lambda: 'NOT_TITLE')
campaigns_by_hide_marker = defaultdict(lambda: True)
for item in db.campaign.find({}, {'guid': True, 'title': True, 'account': True, 'yottosHideSiteMarker': True}):
    campaigns[item.get('guid')] = item.get('title')
    campaigns_by_account[item.get('guid')] = item.get('account')
    campaigns_by_hide_marker[item.get('guid')] = item.get('yottosHideSiteMarker', True)

for item in db.campaign.archive.find({}, {'guid': True, 'title': True, 'account': True, 'yottosHideSiteMarker': True}):
    campaigns[item.get('guid')] = item.get('title')
    campaigns_by_account[item.get('guid')] = item.get('account')
    campaigns_by_hide_marker[item.get('guid')] = item.get('yottosHideSiteMarker', True)

pipeline = [
    {'$match':
        {
            'dt': {'$gte': date_end, '$lt': date_start},
            'adload_cost': {'$gte': 0}
        }
    },
    {'$group':
        {
            '_id': {
                'campaign': '$campaignId',
                'adv': '$inf',
                'offer': '$offer',
                'cost': '$adload_cost',
                "day": {"$dayOfMonth": "$dt"},
                "month": {"$month": "$dt"}
            },
            'url': {'$last': '$url'},
            'count': {
                '$sum': {"$cond": [
                    {"$eq": ["$unique", True]},
                    1,
                    1
                ]
                }
            }
            ,
            'summa': {
                '$sum': {"$cond": [
                    {"$eq": ["$unique", True]},
                    '$adload_cost',
                    0
                ]
                }
            }
        }
    },
    {'$sort': {"_id.day": 1}}
]
data = defaultdict(
    lambda: defaultdict(
        lambda: defaultdict(
            lambda: defaultdict(
                lambda: defaultdict(
                    lambda: defaultdict(
                        lambda: defaultdict(float)
                    )
                )
            )
        )
    )
)
cursor = db.clicks.aggregate(pipeline=pipeline, allowDiskUse=True, useCursor=True)
c = 0
for doc in cursor:
    date = '%s-%s' % (doc['_id']['month'], doc['_id']['day'])
    # date = 'all'
    account = campaigns_by_account[doc['_id']['campaign']]
    campaign = campaigns[doc['_id']['campaign']]
    if campaigns_by_hide_marker[doc['_id']['campaign']]:
        block = doc['_id']['adv']
    else:
        block = informers[doc['_id']['adv']]

    url = doc['url']
    url_parts = list(urlparse.urlparse(url))
    query = dict(urlparse.parse_qsl(url_parts[4]))
    if query.has_key('utm_content'):
        offer = urllib.unquote(query['utm_content'].encode('utf-8'))
    else:
        offer = doc['_id']['offer']
    cost = doc['_id']['cost']
    count = doc['count']
    sum = doc['summa']
    data[account][date][campaign][offer][block][cost]['count'] = count
    data[account][date][campaign][offer][block][cost]['sum'] = sum
    # break

acl = []
for key, value in data.iteritems():  # account
    wb = Workbook()
    sheet_count = 0
    for k, v in value.iteritems():  # date
        ws = wb.create_sheet(k, sheet_count)
        row_count = 1
        ws.cell(row=row_count, column=1).value = 'Название компании'
        ws.cell(row=row_count, column=2).value = 'Id Тизера'
        ws.cell(row=row_count, column=3).value = 'Рекламный блок'
        ws.cell(row=row_count, column=4).value = 'Цена'
        ws.cell(row=row_count, column=5).value = 'Количество кликов'
        ws.cell(row=row_count, column=6).value = 'Сумма'
        c1 = ws['A1']
        c1.font = Font(size=14, bold=True)
        c1.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c1.column].width = 30
        c2 = ws['B1']
        c2.font = Font(size=14, bold=True)
        c2.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c2.column].width = 30
        c3 = ws['C1']
        c3.font = Font(size=14, bold=True)
        c3.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c3.column].width = 38
        c4 = ws['D1']
        c4.font = Font(size=14, bold=True)
        c4.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c4.column].width = 15
        c5 = ws['E1']
        c5.font = Font(size=14, bold=True)
        c5.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c5.column].width = 30
        c6 = ws['F1']
        c6.font = Font(size=14, bold=True)
        c6.alignment = Alignment(shrink_to_fit=False)
        ws.column_dimensions[c6.column].width = 15
        ws.row_dimensions[1].height = 25
        for z, x in v.iteritems():  # campaign
            for a, s in x.iteritems():  # offer
                for q, w in s.iteritems():  # block
                    for e, r in w.iteritems():  # cost
                        row_count += 1
                        ws.cell(row=row_count, column=1).value = z
                        ws.cell(row=row_count, column=2).value = a
                        ws.cell(row=row_count, column=3).value = q
                        ws.cell(row=row_count, column=4).value = e
                        ws.cell(row=row_count, column=5).value = r['count']
                        ws.cell(row=row_count, column=6).value = r['sum']
        sheet_count += 1
    store_dir = os.path.join(rep_dir, '%s-%s-%s' % (date_start.year, date_start.month, date_start.day))
    if not os.path.exists(store_dir):
        os.makedirs(store_dir)
    wb.save(os.path.join(store_dir, str(accounts[key]) + '_%s-%s-%s' % (date_start.year, date_start.month, date_start.day) + '.xlsx'))
