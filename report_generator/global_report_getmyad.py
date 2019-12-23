# -*- coding: UTF-8 -*-
from __future__ import absolute_import, unicode_literals

__author__ = 'kuzmenko-pavel'

import sys
from collections import defaultdict, OrderedDict
from sqlalchemy.orm import joinedload
from datetime import datetime, timedelta
from pytz import timezone
from openpyxl import Workbook
from os.path import join, exists
from os import makedirs

from .models.money.types import MoneyType
from .models.money.defaults import to_money
from .models import ParentAccount
from .models.choiceTypes import ProjectType, AccountType

reload(sys)  # Reload does the trick!
sys.setdefaultencoding('UTF8')


def generate_global_report_getmyad(db_session, store_dir, rates):
    store_dir = join(store_dir, 'getmyad')
    if not exists(store_dir):
        makedirs(store_dir)
    rate = rates[MoneyType.uah]
    now = datetime.now(timezone('Europe/Kiev'))
    now = datetime(now.year, now.month, now.day, 0, 0)
    now = now - timedelta(days=1)
    for account in db_session.query(ParentAccount).options(
            joinedload('statistic'),
            joinedload('statistic_array'),
            joinedload('sites'),
            joinedload('blocks'),
    ).filter(ParentAccount.project == ProjectType.Getmyad,
             ParentAccount.account_type == AccountType.Customer).all():
        data = {
            'account': {
                'name': '',
                'statistic_by_day': defaultdict(lambda: list()),
                'statistic_by_group': defaultdict(lambda: list())
            },
            'site': defaultdict(lambda: {
                'name': '',
                'statistic_by_day': defaultdict(lambda: list()),
                'statistic_by_group': defaultdict(lambda: list())
            }),
            'block': defaultdict(lambda: {
                'name': '',
                'statistic_by_day': defaultdict(lambda: list()),
                'statistic_by_group': defaultdict(lambda: list())
            }),
        }
        account_name = account.name

        data['account']['name'] = account_name
        array_impressions_block = getattr(account.statistic_array, 'array_impressions_block', None)
        array_impressions_block_valid = getattr(account.statistic_array, 'array_impressions_block_valid', None)
        array_clicks_paid = getattr(account.statistic_array, 'array_clicks_paid', None)
        array_clicks_cost = getattr(account.statistic_array, 'array_clicks_cost', None)

        if isinstance(array_impressions_block, list):
            array_impressions_block = array_impressions_block + [0 for x in range(0, 90 - len(array_impressions_block))]
        else:
            array_impressions_block = [0 for x in range(0, 90)]

        if isinstance(array_impressions_block_valid, list):
            array_impressions_block_valid = array_impressions_block_valid + [0 for x in range(0, 90 - len(
                array_impressions_block_valid))]
        else:
            array_impressions_block_valid = [0 for x in range(0, 90)]

        if isinstance(array_clicks_paid, list):
            array_clicks_paid = array_clicks_paid + [0 for x in range(0, 90 - len(array_clicks_paid))]
        else:
            array_clicks_paid = [0 for x in range(0, 90)]

        if isinstance(array_clicks_cost, list):
            array_clicks_cost = array_clicks_cost + [0 for x in range(0, 90 - len(array_clicks_cost))]
        else:
            array_clicks_cost = [0 for x in range(0, 90)]

        for ix, d in enumerate([now - timedelta(days=x) for x in range(0, 90)]):
            data['account']['statistic_by_day'][d] = [array_impressions_block[ix],
                                                      array_impressions_block_valid[ix],
                                                      array_clicks_paid[ix],
                                                      to_money(array_clicks_cost[ix], rate)]

        for site in account.sites:
            id_site = site.id
            data['site'][id_site]['name'] = site.name
            array_impressions_block = getattr(site.statistic_array, 'array_impressions_block', None)
            array_impressions_block_valid = getattr(site.statistic_array, 'array_impressions_block_valid', None)
            array_clicks_paid = getattr(site.statistic_array, 'array_clicks_paid', None)
            array_clicks_cost = getattr(site.statistic_array, 'array_clicks_cost', None)

            if isinstance(array_impressions_block, list):
                array_impressions_block = array_impressions_block + [0 for x in
                                                                     range(0, 90 - len(array_impressions_block))]
            else:
                array_impressions_block = [0 for x in range(0, 90)]

            if isinstance(array_impressions_block_valid, list):
                array_impressions_block_valid = array_impressions_block_valid + [0 for x in range(0, 90 - len(
                    array_impressions_block_valid))]
            else:
                array_impressions_block_valid = [0 for x in range(0, 90)]

            if isinstance(array_clicks_paid, list):
                array_clicks_paid = array_clicks_paid + [0 for x in range(0, 90 - len(array_clicks_paid))]
            else:
                array_clicks_paid = [0 for x in range(0, 90)]

            if isinstance(array_clicks_cost, list):
                array_clicks_cost = array_clicks_cost + [0 for x in range(0, 90 - len(array_clicks_cost))]
            else:
                array_clicks_cost = [0 for x in range(0, 90)]

            for ix, d in enumerate([now - timedelta(days=x) for x in range(0, 90)]):
                data['site'][id_site]['statistic_by_day'][d] = [array_impressions_block[ix],
                                                                array_impressions_block_valid[ix],
                                                                array_clicks_paid[ix],
                                                                to_money(array_clicks_cost[ix], rate)]

        for block in account.blocks:
            id_block = block.id
            data['block'][id_block]['name'] = block.name
            array_impressions_block = getattr(block.statistic_array, 'array_impressions_block', None)
            array_impressions_block_valid = getattr(block.statistic_array, 'array_impressions_block_valid', None)
            array_clicks_paid = getattr(block.statistic_array, 'array_clicks_paid', None)
            array_clicks_cost = getattr(block.statistic_array, 'array_clicks_cost', None)

            if isinstance(array_impressions_block, list):
                array_impressions_block = array_impressions_block + [0 for x in
                                                                     range(0, 90 - len(array_impressions_block))]
            else:
                array_impressions_block = [0 for x in range(0, 90)]

            if isinstance(array_impressions_block_valid, list):
                array_impressions_block_valid = array_impressions_block_valid + [0 for x in range(0, 90 - len(
                    array_impressions_block_valid))]
            else:
                array_impressions_block_valid = [0 for x in range(0, 90)]

            if isinstance(array_clicks_paid, list):
                array_clicks_paid = array_clicks_paid + [0 for x in range(0, 90 - len(array_clicks_paid))]
            else:
                array_clicks_paid = [0 for x in range(0, 90)]

            if isinstance(array_clicks_cost, list):
                array_clicks_cost = array_clicks_cost + [0 for x in range(0, 90 - len(array_clicks_cost))]
            else:
                array_clicks_cost = [0 for x in range(0, 90)]

            for ix, d in enumerate([now - timedelta(days=x) for x in range(0, 90)]):
                data['block'][id_block]['statistic_by_day'][d] = [array_impressions_block[ix],
                                                                  array_impressions_block_valid[ix],
                                                                  array_clicks_paid[ix],
                                                                  to_money(array_clicks_cost[ix], rate)]

        wb = Workbook()
        sheet_count = 0
        row_count = 1
        ws = wb.create_sheet('Аккаунт - %s' % account.name, sheet_count)
        ws.cell(row=row_count, column=1).value = 'ДАТА'
        ws.cell(row=row_count, column=2).value = 'Показы'
        ws.cell(row=row_count, column=3).value = 'Гаранты'
        ws.cell(row=row_count, column=4).value = 'Клики'
        ws.cell(row=row_count, column=5).value = 'Доход'
        statistic_by_day = OrderedDict(sorted(data['account']['statistic_by_day'].items(), reverse=True))
        for z, x in statistic_by_day.iteritems():
            row_count += 1
            ws.cell(row=row_count, column=1).value = z.strftime("%d.%m.%Y")
            ws.cell(row=row_count, column=2).value = int(x[0])
            ws.cell(row=row_count, column=3).value = int(x[1])
            ws.cell(row=row_count, column=4).value = int(x[2])
            ws.cell(row=row_count, column=5).value = x[3]

        sheet_count += 1
        for s in data['site'].itervalues():
            row_count = 1
            ws = wb.create_sheet('Сайт - %s' % s['name'], sheet_count)
            ws.cell(row=row_count, column=1).value = 'ДАТА'
            ws.cell(row=row_count, column=2).value = 'Показы'
            ws.cell(row=row_count, column=3).value = 'Гаранты'
            ws.cell(row=row_count, column=4).value = 'Клики'
            ws.cell(row=row_count, column=5).value = 'Доход'
            statistic_by_day = OrderedDict(sorted(s['statistic_by_day'].items(), reverse=True))
            for z, x in statistic_by_day.iteritems():
                row_count += 1
                ws.cell(row=row_count, column=1).value = z.strftime("%d.%m.%Y")
                ws.cell(row=row_count, column=2).value = int(x[0])
                ws.cell(row=row_count, column=3).value = int(x[1])
                ws.cell(row=row_count, column=4).value = int(x[2])
                ws.cell(row=row_count, column=5).value = x[3]

            sheet_count += 1

        for b in data['block'].itervalues():
            row_count = 1
            ws = wb.create_sheet('Блок - %s' % b['name'], sheet_count)
            ws.cell(row=row_count, column=1).value = 'ДАТА'
            ws.cell(row=row_count, column=2).value = 'Показы'
            ws.cell(row=row_count, column=3).value = 'Гаранты'
            ws.cell(row=row_count, column=4).value = 'Клики'
            ws.cell(row=row_count, column=5).value = 'Доход'
            statistic_by_day = OrderedDict(sorted(b['statistic_by_day'].items(), reverse=True))
            for z, x in statistic_by_day.iteritems():
                row_count += 1
                ws.cell(row=row_count, column=1).value = z.strftime("%d.%m.%Y")
                ws.cell(row=row_count, column=2).value = int(x[0])
                ws.cell(row=row_count, column=3).value = int(x[1])
                ws.cell(row=row_count, column=4).value = int(x[2])
                ws.cell(row=row_count, column=5).value = x[3]
            sheet_count += 1

        wb.save(join(store_dir, account.name + '_%s-%s-%s' % (now.year, now.month, now.day) + '.xlsx'))
