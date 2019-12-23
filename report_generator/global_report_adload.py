# -*- coding: UTF-8 -*-
from __future__ import absolute_import, unicode_literals

__author__ = 'kuzmenko-pavel'

import sys
from collections import defaultdict, OrderedDict
from sqlalchemy.orm import joinedload
from datetime import datetime, timedelta
from pytz import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from os.path import join, exists
from os import makedirs

from .models.money.types import MoneyType
from .models.money.defaults import to_money
from .models import ParentAccount, ParentCampaign, ParentCampaignStatistic, ParentCampaignStatisticArray
from .models.choiceTypes import ProjectType, AccountType, CampaignType

reload(sys)  # Reload does the trick!
sys.setdefaultencoding('UTF8')


def generate_global_report_adload(db_session, store_dir, rates):
    store_dir = join(store_dir, 'adload')
    if not exists(store_dir):
        makedirs(store_dir)
    rate = rates[MoneyType.uah]
    now = datetime.now(timezone('Europe/Kiev'))
    now = datetime(now.year, now.month, now.day, 0, 0)
    now = now - timedelta(days=1)
    for account in db_session.query(ParentAccount).filter(ParentAccount.project == ProjectType.Adload,
                                                          ParentAccount.account_type == AccountType.Customer).all():

        data = {
            'new_auditory': defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(0.0)))),
            'remarketing': defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(0.0)))),
            'relevant_auditory': defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(0.0)))),
        }

        campaigns = db_session.query(ParentCampaign).join(
            ParentCampaignStatistic, ParentCampaignStatistic.id == ParentCampaign.id, isouter=True
        ).join(
            ParentCampaignStatisticArray, ParentCampaignStatisticArray.id == ParentCampaign.id, isouter=True
        ).options(
            joinedload('statistic'),
            joinedload('statistic_array'),
        ).filter(ParentCampaign.id_account == account.id,
                 ParentCampaign.campaign_type.in_([CampaignType.new_auditory, CampaignType.relevant_auditory,
                                                   CampaignType.remarketing])).all()
        new_auditory_processing = False
        remarketing_processing = False
        relevant_auditory_processing = False
        for campaign in campaigns:
            if campaign.campaign_type == CampaignType.new_auditory:
                if campaign.statistic_array:
                    new_auditory_processing = True
                    clicks_paid_array = getattr(campaign.statistic_array, 'array_clicks_paid', None)
                    if isinstance(clicks_paid_array, list):
                        clicks_paid_array = clicks_paid_array + [0 for x in range(0, 90 - len(clicks_paid_array))]
                    else:
                        clicks_paid_array = [0 for x in range(0, 90)]

                    clicks_cost_array = getattr(campaign.statistic_array, 'array_clicks_cost', None)
                    if isinstance(clicks_cost_array, list):
                        clicks_cost_array = clicks_cost_array + [0 for x in range(0, 90 - len(clicks_cost_array))]
                    else:
                        clicks_cost_array = [0 for x in range(0, 90)]

                    for ix, date in enumerate([now - timedelta(days=x) for x in range(0, 90)]):
                        data['new_auditory'][date][campaign.name]['count'] = clicks_paid_array[ix]
                        data['new_auditory'][date][campaign.name]['summ'] = to_money(clicks_cost_array[ix], rate)

            elif campaign.campaign_type == CampaignType.relevant_auditory:
                if campaign.statistic_array:
                    relevant_auditory_processing = True
                    clicks_paid_array = getattr(campaign.statistic_array, 'array_clicks_paid', None)
                    if isinstance(clicks_paid_array, list):
                        clicks_paid_array = clicks_paid_array + [0 for x in range(0, 90 - len(clicks_paid_array))]
                    else:
                        clicks_paid_array = [0 for x in range(0, 90)]

                    clicks_cost_array = getattr(campaign.statistic_array, 'array_clicks_cost', None)
                    if isinstance(clicks_cost_array, list):
                        clicks_cost_array = clicks_cost_array + [0 for x in range(0, 90 - len(clicks_cost_array))]
                    else:
                        clicks_cost_array = [0 for x in range(0, 90)]

                    for ix, date in enumerate([now - timedelta(days=x) for x in range(0, 90)]):
                        data['relevant_auditory'][date][campaign.name]['count'] = clicks_paid_array[ix]
                        data['relevant_auditory'][date][campaign.name]['summ'] = to_money(clicks_cost_array[ix], rate)
            elif campaign.campaign_type == CampaignType.remarketing:
                if campaign.statistic_array:
                    remarketing_processing = True
                    clicks_paid_array = getattr(campaign.statistic_array, 'array_clicks_paid', None)
                    if isinstance(clicks_paid_array, list):
                        clicks_paid_array = clicks_paid_array + [0 for x in range(0, 90 - len(clicks_paid_array))]
                    else:
                        clicks_paid_array = [0 for x in range(0, 90)]

                    clicks_cost_array = getattr(campaign.statistic_array, 'array_clicks_cost', None)
                    if isinstance(clicks_cost_array, list):
                        clicks_cost_array = clicks_cost_array + [0 for x in range(0, 90 - len(clicks_cost_array))]
                    else:
                        clicks_cost_array = [0 for x in range(0, 90)]

                    for ix, date in enumerate([now - timedelta(days=x) for x in range(0, 90)]):
                        data['remarketing'][date][campaign.name]['count'] = clicks_paid_array[ix]
                        data['remarketing'][date][campaign.name]['summ'] = to_money(clicks_cost_array[ix], rate)

        wb = Workbook()
        sheet_count = 0
        if new_auditory_processing:
            ws = wb.create_sheet('Новая аудитория ', sheet_count)
            ws.cell(row=1, column=1).value = 'ДАТА'
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            new_auditory = OrderedDict(sorted(data['new_auditory'].items(), reverse=True))
            row_count = 2
            for k, v in new_auditory.iteritems():
                column_count = 2
                row_count += 1
                ws.cell(row=row_count, column=1).value = k.strftime("%d.%m.%Y")
                for z, x in v.iteritems():
                    ws.cell(row=1, column=column_count).value = z
                    ws.merge_cells(start_row=1, start_column=column_count, end_row=1, end_column=column_count + 1)
                    ws.cell(row=2, column=column_count).value = 'клики'
                    ws.cell(row=2, column=column_count + 1).value = 'сумма'
                    ws.cell(row=row_count, column=column_count).value = x.get('count')
                    ws.cell(row=row_count, column=column_count + 1).value = x.get('summ')
                    column_count += 2
                column_letter_start = get_column_letter(2)
                column_letter_stop = get_column_letter(column_count - 2)
                # ws.cell(row=row_count, column=column_count).value = "=SUM(%s%s:%s%s)" % (column_letter_start,
                #                                                                           row_count,
                #                                                                           column_letter_stop,
                #                                                                           row_count)
            ws.cell(row=1, column=column_count).value = 'Всего'
            ws.merge_cells(start_row=1, start_column=column_count, end_row=2, end_column=column_count)
            ws.cell(row=row_count + 1, column=1).value = 'Всего'
            for x in range(2, column_count):
                column_letter = get_column_letter(x)
                ws.cell(row=row_count + 1, column=x).value = "=SUM(%s%s:%s%s)" % (column_letter,
                                                                                   3,
                                                                                   column_letter,
                                                                                   row_count)

            sheet_count += 1
        if remarketing_processing:
            ws = wb.create_sheet('Ремаркетинг ', sheet_count)
            ws.cell(row=1, column=1).value = 'ДАТА'
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            remarketing = OrderedDict(sorted(data['remarketing'].items(), reverse=True))
            row_count = 2
            for k, v in remarketing.iteritems():
                column_count = 2
                row_count += 1
                ws.cell(row=row_count, column=1).value = k.strftime("%d.%m.%Y")
                for z, x in v.iteritems():
                    ws.cell(row=1, column=column_count).value = z
                    ws.merge_cells(start_row=1, start_column=column_count, end_row=1, end_column=column_count + 1)
                    ws.cell(row=2, column=column_count).value = 'клики'
                    ws.cell(row=2, column=column_count + 1).value = 'сумма'
                    ws.cell(row=row_count, column=column_count).value = x.get('count')
                    ws.cell(row=row_count, column=column_count + 1).value = x.get('summ')
                    column_count += 2
                column_letter_start = get_column_letter(2)
                column_letter_stop = get_column_letter(column_count - 2)
                # ws.cell(row=row_count, column=column_count).value = "=SUM(%s%s:%s%s)" % (column_letter_start,
                #                                                                           row_count,
                #                                                                           column_letter_stop,
                #                                                                           row_count)
            ws.cell(row=1, column=column_count).value = 'Всего'
            ws.merge_cells(start_row=1, start_column=column_count, end_row=2, end_column=column_count)
            ws.cell(row=row_count + 1, column=1).value = 'Всего'
            for x in range(2, column_count):
                column_letter = get_column_letter(x)
                ws.cell(row=row_count + 1, column=x).value = "=SUM(%s%s:%s%s)" % (column_letter,
                                                                                   3,
                                                                                   column_letter,
                                                                                   row_count)

            sheet_count += 1
        if relevant_auditory_processing:
            ws = wb.create_sheet('Релевантная ', sheet_count)
            ws.cell(row=1, column=1).value = 'ДАТА'
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            relevant_auditory = OrderedDict(sorted(data['relevant_auditory'].items(), reverse=True))
            row_count = 2
            for k, v in relevant_auditory.iteritems():
                column_count = 2
                row_count += 1
                ws.cell(row=row_count, column=1).value = k.strftime("%d.%m.%Y")
                for z, x in v.iteritems():
                    ws.cell(row=1, column=column_count).value = z
                    ws.merge_cells(start_row=1, start_column=column_count, end_row=1, end_column=column_count + 1)
                    ws.cell(row=2, column=column_count).value = 'клики'
                    ws.cell(row=2, column=column_count + 1).value = 'сумма'
                    ws.cell(row=row_count, column=column_count).value = x.get('count')
                    ws.cell(row=row_count, column=column_count + 1).value = x.get('summ')
                    column_count += 2
                column_letter_start = get_column_letter(2)
                column_letter_stop = get_column_letter(column_count - 2)
                # ws.cell(row=row_count, column=column_count).value = "=SUM(%s%s:%s%s)" % (column_letter_start,
                #                                                                           row_count,
                #                                                                           column_letter_stop,
                #                                                                           row_count)
            ws.cell(row=1, column=column_count).value = 'Всего'
            ws.merge_cells(start_row=1, start_column=column_count, end_row=2, end_column=column_count)
            ws.cell(row=row_count + 1, column=1).value = 'Всего'
            for x in range(2, column_count):
                column_letter = get_column_letter(x)
                ws.cell(row=row_count + 1, column=x).value = "=SUM(%s%s:%s%s)" % (column_letter,
                                                                                   3,
                                                                                   column_letter,
                                                                                   row_count)

        if any([new_auditory_processing, remarketing_processing, relevant_auditory_processing]):
            wb.save(join(store_dir, account.name + '_%s-%s-%s' % (now.year, now.month, now.day) + '.xlsx'))
