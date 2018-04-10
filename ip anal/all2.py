# -*- coding: utf-8 -*-
from pymongo import MongoClient
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

main_db_host = 'srv-5.yottos.com:27018,srv-5.yottos.com:27019,srv-5.yottos.com:27020'
conn = MongoClient(host=main_db_host)
db = conn.getmyad_db


def calc(db, dateS, dateE):
    print('%s - %s' % (dateS, dateE))

    pipeline = [
        {'$match':
             {'date': {'$gte': dateS, '$lt': dateE}, 'impressions': {'$gte': 0}}
         },
        {'$group': {
            '_id': '$ip',
            'impressions': {'$sum': {"$cond": [{"$gt": ["$impressions", 0]}, 1, 0]}},
            'all_clicks': {'$sum': {"$cond": [{"$gt": ["$all_clicks", 0]}, 1, 0]}}
        },
        },
        {'$match':
             {}
         },
        {'$group': {
            '_id': {'impressions': '$impressions', 'all_clicks': '$all_clicks'},
            'count': {'$sum': 1}
        },
        }
    ]

    cursor = db.ip.stats.daily.raw.aggregate(pipeline=pipeline, cursor={}, allowDiskUse=True)
    wb = Workbook()
    ws = wb.create_sheet('Mysheet')
    for doc in cursor:
        row = doc['_id']['all_clicks']
        column = doc['_id']['impressions']
        value = doc['count']
        ws.cell(row=row+1, column=column+1).value = value
    wb.save('rep.xlsx')


dateS = datetime.datetime(2018, 02, 27, 0, 0)
dateE = datetime.datetime(2018, 04, 04, 0, 0)
calc(db, dateS, dateE)